import argparse
import csv
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dotenv import load_dotenv

load_dotenv()

from app import db  # noqa: E402

DEFAULT_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "app",
    "exam_content",
    "Claude_Architect_Foundations_Master_Question_List.xlsx",
)

DEFAULT_QUESTION_SET_NAME = "Claude Architect Foundations Master Question List"

# Column order, positional, for both the .xlsx and .csv input formats:
# Module, Question, Answer Choice 1, Answer Choice 2, Answer Choice 3,
# Answer Choice 4, Correct Choice Number, Correct Choice, Justification
#
# Correct Choice Number/Correct Choice are usually single-valued (single-
# answer question), but may instead describe a multi-select ("select all
# that apply") question with more than one correct answer, in either of two
# conventions seen across real content files:
#   - ";"-separated: "1;3;4" / "Choice A; Choice C; Choice D" (one Correct
#     Choice part per index, exact-matching the corresponding Answer Choice).
#   - ","-separated: "1,3" / "1. Choice A\n3. Choice C" (Correct Choice is a
#     single field with one "N. <text>" line per index, joined by newlines).


def _split_correct_desc(correct_desc, delimiter):
    if delimiter == ",":
        parts = str(correct_desc).split("\n")
        return [re.sub(r"^\d+\.\s*", "", p.strip()) for p in parts]
    return [p.strip() for p in str(correct_desc).split(delimiter)]


def _parse_row(row_num, row):
    module, question, c1, c2, c3, c4, correct_num, correct_desc, justification = row[:9]

    fields = {
        "Module": module,
        "Question": question,
        "Answer Choice 1": c1,
        "Answer Choice 2": c2,
        "Answer Choice 3": c3,
        "Answer Choice 4": c4,
        "Justification": justification,
    }
    errors = []
    for label, value in fields.items():
        if value is None or str(value).strip() == "":
            errors.append(f"Row {row_num}: '{label}' is blank.")

    choices = [c1, c2, c3, c4]

    delimiter = ";" if ";" in str(correct_num) else ","
    correct_indices = []
    for part in str(correct_num).split(delimiter):
        part = part.strip()
        try:
            idx = int(part) - 1
        except (TypeError, ValueError):
            errors.append(f"Row {row_num}: 'Correct Choice Number' ({correct_num!r}) is not a number.")
            return None, errors
        if not (0 <= idx <= 3):
            errors.append(f"Row {row_num}: 'Correct Choice Number' ({correct_num!r}) is not between 1 and 4.")
            return None, errors
        correct_indices.append(idx)

    if len(set(correct_indices)) != len(correct_indices):
        errors.append(f"Row {row_num}: 'Correct Choice Number' ({correct_num!r}) has a repeated index.")
        return None, errors

    if errors:
        return None, errors

    # Only split "Correct Choice" for genuinely multi-value rows — a single-
    # answer row's text may itself legitimately contain a literal ";" or ","
    # (e.g. "...among several; data quality..."). Some source files prefix
    # even single-answer rows with "N. " (matching the multi-select "N. "
    # convention) — strip it the same way; a no-op when it's not present.
    if len(correct_indices) == 1:
        correct_desc_parts = [re.sub(r"^\d+\.\s*", "", str(correct_desc).strip())]
    else:
        correct_desc_parts = _split_correct_desc(correct_desc, delimiter)
    if len(correct_desc_parts) != len(correct_indices):
        errors.append(
            f"Row {row_num}: 'Correct Choice' has {len(correct_desc_parts)} part(s) but "
            f"'Correct Choice Number' ({correct_num!r}) has {len(correct_indices)}."
        )
        return None, errors

    for idx, desc_part in zip(correct_indices, correct_desc_parts):
        if str(choices[idx]).strip() != desc_part:
            errors.append(
                f"Row {row_num}: 'Correct Choice' text does not match Answer Choice {idx + 1}."
            )
            return None, errors

    return {
        "module": str(module).strip(),
        "question_text": str(question).strip(),
        "options": [str(c).strip() for c in choices],
        "correct_indices": correct_indices,
        "is_multi_select": len(correct_indices) > 1,
        "justification": str(justification).strip(),
    }, []


def _is_blank_row(row):
    return row is None or all(c is None or str(c).strip() == "" for c in row)


def _load_from_xlsx(path):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Master Question List"] if "Master Question List" in wb.sheetnames else wb.worksheets[0]

    questions = []
    errors = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if _is_blank_row(row):
            continue
        question, row_errors = _parse_row(row_num, row)
        if row_errors:
            errors.extend(row_errors)
            continue
        questions.append(question)

    return questions, errors


def _load_from_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))

    questions = []
    errors = []
    for row_num, row in enumerate(rows[1:], start=2):  # row 1 is the header
        if _is_blank_row(row):
            continue
        if len(row) < 9:
            errors.append(f"Row {row_num}: expected 9 columns, found {len(row)}.")
            continue
        question, row_errors = _parse_row(row_num, row)
        if row_errors:
            errors.extend(row_errors)
            continue
        questions.append(question)

    return questions, errors


def _get_or_create_question_set(name):
    rs = db.execute("SELECT id FROM question_sets WHERE name = ?", [name])
    if rs.rows:
        return rs.rows[0][0]
    rs = db.execute("INSERT INTO question_sets (name) VALUES (?)", [name])
    return rs.last_insert_rowid


def _get_or_create_module(name, order_index, cache):
    if name in cache:
        return cache[name]
    rs = db.execute("SELECT id FROM exam_modules WHERE name = ?", [name])
    if rs.rows:
        module_id = rs.rows[0][0]
    else:
        rs = db.execute(
            "INSERT INTO exam_modules (name, order_index) VALUES (?, ?)",
            [name, order_index],
        )
        module_id = rs.last_insert_rowid
    cache[name] = module_id
    return module_id


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Import questions from a .xlsx or .csv file with columns: Module, Question, "
            "Answer Choice 1-4, Correct Choice Number, Correct Choice, Justification."
        )
    )
    parser.add_argument("path", nargs="?", default=DEFAULT_PATH)
    parser.add_argument(
        "--set-name",
        default=DEFAULT_QUESTION_SET_NAME,
        help="Question set to load into (created if it doesn't exist yet).",
    )
    args = parser.parse_args()

    ext = os.path.splitext(args.path)[1].lower()
    if ext == ".xlsx":
        questions, errors = _load_from_xlsx(args.path)
    elif ext == ".csv":
        questions, errors = _load_from_csv(args.path)
    else:
        print(f"Unsupported file type: {ext} (use .xlsx or .csv)")
        sys.exit(1)

    if errors:
        print(f"Found {len(errors)} problem(s) — nothing was written to the database:")
        for err in errors:
            print(f"  - {err}")
        sys.exit(1)

    if not questions:
        print("No questions found in the file.")
        sys.exit(1)

    set_id = _get_or_create_question_set(args.set_name)

    module_cache = {}
    module_order = 0
    option_count = 0
    for i, q in enumerate(questions):
        if q["module"] not in module_cache:
            module_id = _get_or_create_module(q["module"], module_order, module_cache)
            module_order += 1
        else:
            module_id = module_cache[q["module"]]

        rs = db.execute(
            "INSERT INTO questions (set_id, module_id, question_text, points, order_index, is_multi_select) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            [set_id, module_id, q["question_text"], 3, i, 1 if q["is_multi_select"] else 0],
        )
        question_id = rs.last_insert_rowid

        option_ids = []
        for j, opt_text in enumerate(q["options"]):
            rs = db.execute(
                "INSERT INTO options (question_id, option_text, order_index, is_correct) "
                "VALUES (?, ?, ?, ?)",
                [question_id, opt_text, j, 1 if j in q["correct_indices"] else 0],
            )
            option_ids.append(rs.last_insert_rowid)
            option_count += 1

        # answer_key only ever stores one correct_option_id (schema constraint
        # predates multi-select); options.is_correct above is authoritative.
        correct_option_id = option_ids[q["correct_indices"][0]]
        db.execute(
            "INSERT INTO answer_key (question_id, correct_option_id, justification) "
            "VALUES (?, ?, ?)",
            [question_id, correct_option_id, q["justification"]],
        )

    print(
        f"Seeded into question set {args.set_name!r}: {len(module_cache)} modules, "
        f"{len(questions)} questions, {option_count} options, {len(questions)} answer-key rows."
    )


if __name__ == "__main__":
    main()
    db.close_client()
